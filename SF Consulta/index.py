import openpyxl

# Caminho absoluto do arquivo
caminho_arquivo = r'C:\Users\edson.dasilva.ext\Documents\github\automacao-trampo\SF Consulta\Consulta.xlsx'

# Carrega a planilha
wb = openpyxl.load_workbook(caminho_arquivo)

# Escolhe a planilha com a qual você está trabalhando
sheet = wb['SF_Consulta']

# Obtém as dimensões da planilha (número de linhas e colunas)
linhas = sheet.max_row
colunas = sheet.max_column

# Exibe todos os valores na planilha no console
for i in range(1, linhas + 1):
    for j in range(1, colunas + 1):
        valor = sheet.cell(row=i, column=j).value
        print(f"Valor na linha {i}, coluna {j}: {valor}")

# Fecha a planilha
wb.close()
