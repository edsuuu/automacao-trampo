import os
import openpyxl
from openpyxl import load_workbook
from tqdm import tqdm

# Caminho para a pasta onde estão as planilhas
pasta_planilhas = r"C:\Users\edson.dasilva.ext\Desktop\docusing\Planilhas"

# Lista todos os arquivos na pasta de planilhas
arquivos_planilhas = os.listdir(pasta_planilhas)

# Inicialize um novo arquivo Excel
arquivo_final = openpyxl.Workbook()
planilha_final = arquivo_final.active

# Use tqdm para criar uma barra de progresso
progresso = tqdm(arquivos_planilhas, desc="Combinando planilhas", unit="planilha")

# Itera sobre os arquivos de planilha na pasta
for arquivo in progresso:
    if arquivo.endswith(".xlsx"):
        caminho_arquivo = os.path.join(pasta_planilhas, arquivo)
        # Abra o arquivo da planilha
        workbook = load_workbook(filename=caminho_arquivo)
        # Selecione a planilha ativa (pode ser necessário ajustar se a planilha não for a primeira)
        planilha = workbook.active
        # Itere pelas linhas da planilha e copie os valores para a planilha final
        for row in planilha.iter_rows(values_only=True):
            planilha_final.append(row)

# Salve a planilha final
caminho_saida = r"C:\Users\edson.dasilva.ext\Desktop\docusing\Planilhas\Planilha_Combinada.xlsx"
arquivo_final.save(caminho_saida)

print(f"Dados combinados salvos em {caminho_saida}")